import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNo, colNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNo, column=colNo).value


def writeData(file, sheetName, rowNo, colNo, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNo, column=colNo).value = data
    workbook.save(file)